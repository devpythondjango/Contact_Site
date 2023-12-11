from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from .views import login_decorator
from contact.models import ApplicationCreate, Application
from dashboard.forms import ApplicationCreateForm
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side


@login_decorator
def export_to_excel_id(request, pk):
    # Olingan ariza
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').get(pk=pk)

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'Jins', 'Hujum turi',
                  'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami',
                  'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova',
                  'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z', 'Status']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data row
    ws['A2'] = f'{application.first_name} {application.last_name} {application.surname}'
    ws['B2'] = application.phone
    ws['C2'] = application.birthday
    ws['D2'] = application.passport_serial
    ws['E2'] = application.district.tuman_name if application.district else ''
    ws['F2'] = application.address
    ws['G2'] = application.gender
    ws['H2'] = application.hujumturi.hujum_name if application.hujumturi else ''
    ws['I2'] = application.plastikraqam_ozi
    ws['J2'] = application.plastikraqam_gumondor
    ws['K2'] = application.full_name_gumondor
    ws['L2'] = application.phone_gumondor
    ws['M2'] = application.vaqt
    ws['N2'] = application.summa
    ws['O2'] = application.ilova
    ws['P2'] = application.ilova_gumondor
    ws['Q2'] = application.text
    ws['R2'] = application.app_create.status if application.app_create else ''

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Excel faylini HTTP javob qilib qaytarish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ariza_{pk}.xlsx'
    response.write(save_virtual_workbook(wb))

    return response

@login_decorator
def ariza_view(request, pk):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').get(pk=pk)
    ctx = {'application': application}
    return render(request, 'dashboard/ariza_view.html', ctx)


@login_decorator
def ariza_edit(request, pk):
    application = Application.objects.get(pk=pk)
    app_create = ApplicationCreate.objects.get(pk=pk)
    application_create = ApplicationCreate.objects.get(application=application)

    form = ApplicationCreateForm(request.POST or None, instance=application_create)

    if request.method == 'POST' and form.is_valid():
        form.save()

        return redirect('tablitsa')

    ctx = {
        "application": application,
        "app_create": app_create,
        "form": form
    }
    return render(request, 'dashboard/ariza_edit.html', ctx)