from django.shortcuts import render
from contact.models import Application, Tuman, HujumTuri
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from openpyxl import Workbook
from .views import login_decorator
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side


@login_decorator
def status_yangi(request):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(
        app_create__status='yangi')

    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    hujumturi_id = request.GET.get('hujum', '')
    district_id = request.GET.get('tuman', '')

    if first_name:
        application = application.filter(first_name__icontains=first_name)
    if last_name:
        application = application.filter(last_name__icontains=last_name)
    if district_id:
        application = application.filter(district_id=district_id)
    if hujumturi_id:
        application = application.filter(hujumturi_id=hujumturi_id)

    tumans = Tuman.objects.all()
    hujumturi = HujumTuri.objects.all()

    paginator = Paginator(application, 10)
    page = request.GET.get('page')

    try:
        applications = paginator.page(page)
    except PageNotAnInteger:
        applications = paginator.page(1)
    except EmptyPage:
        applications = paginator.page(paginator.num_pages)

    ctx = {'applications': applications, 'tumans': tumans, 'hujumturi': hujumturi}
    return render(request, 'dashboard/holatlar/yangi.html', ctx)


@login_decorator
def export_to_excel_status1(request):
    applications = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='yangi')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'jins','Hujum turi', 'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami', 'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova', 'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data rows
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = f'{app.first_name} {app.last_name} {app.surname}'
        ws[f'B{row_num}'] = app.phone
        ws[f'C{row_num}'] = app.birthday
        ws[f'D{row_num}'] = app.passport_serial
        ws[f'E{row_num}'] = app.district.tuman_name if app.district else ''
        ws[f'F{row_num}'] = app.address
        ws[f'G{row_num}'] = app.gender
        ws[f'H{row_num}'] = app.hujumturi.hujum_name if app.hujumturi else ''
        ws[f'I{row_num}'] = app.plastikraqam_ozi
        ws[f'J{row_num}'] = app.plastikraqam_gumondor
        ws[f'K{row_num}'] = app.full_name_gumondor
        ws[f'L{row_num}'] = app.phone_gumondor
        ws[f'M{row_num}'] = app.vaqt
        ws[f'N{row_num}'] = app.summa
        ws[f'O{row_num}'] = app.ilova
        ws[f'P{row_num}'] = app.ilova_gumondor
        ws[f'Q{row_num}'] = app.text

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Create and return the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_status_yangi.xlsx'
    wb.save(response)

    return response


@login_decorator
def status_tekshirilmoqda(request):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='tekshirilmoqda')

    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    hujumturi_id = request.GET.get('hujum', '')
    district_id = request.GET.get('tuman', '')

    if first_name:
        application = application.filter(first_name__icontains=first_name)
    if last_name:
        application = application.filter(last_name__icontains=last_name)
    if district_id:
        application = application.filter(district_id=district_id)
    if hujumturi_id:
        application = application.filter(hujumturi_id=hujumturi_id)

    tumans = Tuman.objects.all()
    hujumturi = HujumTuri.objects.all()

    paginator = Paginator(application, 10)
    page = request.GET.get('page')

    try:
        applications = paginator.page(page)
    except PageNotAnInteger:
        applications = paginator.page(1)
    except EmptyPage:
        applications = paginator.page(paginator.num_pages)

    ctx = {'applications': applications, 'tumans': tumans, 'hujumturi': hujumturi}
    return render(request, 'dashboard/holatlar/tekshirilmoqda.html', ctx)


@login_decorator
def export_to_excel_status2(request):
    applications = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='tekshirilmoqda')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'jins','Hujum turi', 'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami', 'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova', 'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data rows
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = f'{app.first_name} {app.last_name} {app.surname}'
        ws[f'B{row_num}'] = app.phone
        ws[f'C{row_num}'] = app.birthday
        ws[f'D{row_num}'] = app.passport_serial
        ws[f'E{row_num}'] = app.district.tuman_name if app.district else ''
        ws[f'F{row_num}'] = app.address
        ws[f'G{row_num}'] = app.gender
        ws[f'H{row_num}'] = app.hujumturi.hujum_name if app.hujumturi else ''
        ws[f'I{row_num}'] = app.plastikraqam_ozi
        ws[f'J{row_num}'] = app.plastikraqam_gumondor
        ws[f'K{row_num}'] = app.full_name_gumondor
        ws[f'L{row_num}'] = app.phone_gumondor
        ws[f'M{row_num}'] = app.vaqt
        ws[f'N{row_num}'] = app.summa
        ws[f'O{row_num}'] = app.ilova
        ws[f'P{row_num}'] = app.ilova_gumondor
        ws[f'Q{row_num}'] = app.text

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Create and return the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_status_tekshirilmoqda.xlsx'
    wb.save(response)

    return response


@login_decorator
def status_tekshirildi(request):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='tekshirildi')

    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    hujumturi_id = request.GET.get('hujum', '')
    district_id = request.GET.get('tuman', '')

    if first_name:
        application = application.filter(first_name__icontains=first_name)
    if last_name:
        application = application.filter(last_name__icontains=last_name)
    if district_id:
        application = application.filter(district_id=district_id)
    if hujumturi_id:
        application = application.filter(hujumturi_id=hujumturi_id)

    tumans = Tuman.objects.all()
    hujumturi = HujumTuri.objects.all()

    paginator = Paginator(application, 10)
    page = request.GET.get('page')

    try:
        applications = paginator.page(page)
    except PageNotAnInteger:
        applications = paginator.page(1)
    except EmptyPage:
        applications = paginator.page(paginator.num_pages)

    ctx = {'applications': applications, 'tumans': tumans, 'hujumturi': hujumturi}
    return render(request, 'dashboard/holatlar/tekshirildi.html', ctx)


@login_decorator
def export_to_excel_status3(request):
    applications = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='tekshirildi')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'jins','Hujum turi', 'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami', 'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova', 'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data rows
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = f'{app.first_name} {app.last_name} {app.surname}'
        ws[f'B{row_num}'] = app.phone
        ws[f'C{row_num}'] = app.birthday
        ws[f'D{row_num}'] = app.passport_serial
        ws[f'E{row_num}'] = app.district.tuman_name if app.district else ''
        ws[f'F{row_num}'] = app.address
        ws[f'G{row_num}'] = app.gender
        ws[f'H{row_num}'] = app.hujumturi.hujum_name if app.hujumturi else ''
        ws[f'I{row_num}'] = app.plastikraqam_ozi
        ws[f'J{row_num}'] = app.plastikraqam_gumondor
        ws[f'K{row_num}'] = app.full_name_gumondor
        ws[f'L{row_num}'] = app.phone_gumondor
        ws[f'M{row_num}'] = app.vaqt
        ws[f'N{row_num}'] = app.summa
        ws[f'O{row_num}'] = app.ilova
        ws[f'P{row_num}'] = app.ilova_gumondor
        ws[f'Q{row_num}'] = app.text

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Create and return the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_status_tekshirildi.xlsx'
    wb.save(response)

    return response


@login_decorator
def status_rad_etildi(request):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='rad etildi')

    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    hujumturi_id = request.GET.get('hujum', '')
    district_id = request.GET.get('tuman', '')

    if first_name:
        application = application.filter(first_name__icontains=first_name)
    if last_name:
        application = application.filter(last_name__icontains=last_name)
    if district_id:
        application = application.filter(district_id=district_id)
    if hujumturi_id:
        application = application.filter(hujumturi_id=hujumturi_id)

    tumans = Tuman.objects.all()
    hujumturi = HujumTuri.objects.all()

    paginator = Paginator(application, 10)
    page = request.GET.get('page')

    try:
        applications = paginator.page(page)
    except PageNotAnInteger:
        applications = paginator.page(1)
    except EmptyPage:
        applications = paginator.page(paginator.num_pages)

    ctx = {'applications': applications, 'tumans': tumans, 'hujumturi': hujumturi}
    return render(request, 'dashboard/holatlar/rad_etildi.html', ctx)


@login_decorator
def export_to_excel_status4(request):
    applications = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='rad etildi')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'jins','Hujum turi', 'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami', 'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova', 'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data rows
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = f'{app.first_name} {app.last_name} {app.surname}'
        ws[f'B{row_num}'] = app.phone
        ws[f'C{row_num}'] = app.birthday
        ws[f'D{row_num}'] = app.passport_serial
        ws[f'E{row_num}'] = app.district.tuman_name if app.district else ''
        ws[f'F{row_num}'] = app.address
        ws[f'G{row_num}'] = app.gender
        ws[f'H{row_num}'] = app.hujumturi.hujum_name if app.hujumturi else ''
        ws[f'I{row_num}'] = app.plastikraqam_ozi
        ws[f'J{row_num}'] = app.plastikraqam_gumondor
        ws[f'K{row_num}'] = app.full_name_gumondor
        ws[f'L{row_num}'] = app.phone_gumondor
        ws[f'M{row_num}'] = app.vaqt
        ws[f'N{row_num}'] = app.summa
        ws[f'O{row_num}'] = app.ilova
        ws[f'P{row_num}'] = app.ilova_gumondor
        ws[f'Q{row_num}'] = app.text

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Create and return the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_status_rad_etildi.xlsx'
    wb.save(response)

    return response


@login_decorator
def status_malumot_topilmadi(request):
    application = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='ma\'lumot topilmadi')

    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    hujumturi_id = request.GET.get('hujum', '')
    district_id = request.GET.get('tuman', '')

    if first_name:
        application = application.filter(first_name__icontains=first_name)
    if last_name:
        application = application.filter(last_name__icontains=last_name)
    if district_id:
        application = application.filter(district_id=district_id)
    if hujumturi_id:
        application = application.filter(hujumturi_id=hujumturi_id)

    tumans = Tuman.objects.all()
    hujumturi = HujumTuri.objects.all()

    paginator = Paginator(application, 10)
    page = request.GET.get('page')

    try:
        applications = paginator.page(page)
    except PageNotAnInteger:
        applications = paginator.page(1)
    except EmptyPage:
        applications = paginator.page(paginator.num_pages)

    ctx = {'applications': applications, 'tumans': tumans, 'hujumturi': hujumturi}
    return render(request, 'dashboard/holatlar/malumot_topilmadi.html', ctx)


@login_decorator
def export_to_excel_status5(request):
    applications = Application.objects.select_related('hujumturi', 'district', 'app_create').filter(app_create__status='ma\'lumot topilmadi')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_style = Font(bold=True, color='FFFFFF')
    cell_style = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='0072BA', end_color='0072BA', fill_type='solid')

    # Header row
    header_row = ['F.I.Sh', 'Telefon', 'Tug\'ilgan yili', 'Passport', 'Tuman', 'Manzil', 'jins','Hujum turi', 'Plastik raqami', 'Plastik raqami \ngumondor', 'Gumondor F.I.Sh', 'Gumondor \n telefon raqami', 'Pul yechilgan \nsana', 'Pul yechib \nolingan summa', 'Shaxs ishlatgan\n ilova', 'Gumondor\n ishlatgan ilova', 'Shaxs yozgan\n qisqa so\'z']
    for col_num, value in enumerate(header_row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = value
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = cell_style
        ws[f'{col_letter}1'].fill = header_fill

    # Data rows
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = f'{app.first_name} {app.last_name} {app.surname}'
        ws[f'B{row_num}'] = app.phone
        ws[f'C{row_num}'] = app.birthday
        ws[f'D{row_num}'] = app.passport_serial
        ws[f'E{row_num}'] = app.district.tuman_name if app.district else ''
        ws[f'F{row_num}'] = app.address
        ws[f'G{row_num}'] = app.gender
        ws[f'H{row_num}'] = app.hujumturi.hujum_name if app.hujumturi else ''
        ws[f'I{row_num}'] = app.plastikraqam_ozi
        ws[f'J{row_num}'] = app.plastikraqam_gumondor
        ws[f'K{row_num}'] = app.full_name_gumondor
        ws[f'L{row_num}'] = app.phone_gumondor
        ws[f'M{row_num}'] = app.vaqt
        ws[f'N{row_num}'] = app.summa
        ws[f'O{row_num}'] = app.ilova
        ws[f'P{row_num}'] = app.ilova_gumondor
        ws[f'Q{row_num}'] = app.text

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    # Create and return the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_status_malumot_topilmadi.xlsx'
    wb.save(response)

    return response
